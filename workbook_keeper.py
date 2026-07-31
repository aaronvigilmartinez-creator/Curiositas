from openpyxl import load_workbook

workbook = load_workbook("GPK_Original_Series_Through_Flashback3_Checklist.xlsx")

sheet = workbook["OS1"]
beacons = []

for row in sheet.iter_rows(values_only=True):

    if (
    isinstance(row[1], str)
    and len(row[1]) > 1
    and row[1][-1] in ("a", "b")
):

        beacon = {

    "universe": "Garbage Pail Kids",

    "series": "Original Series 1",

    "worksheet": sheet.title,

    "a_number": row[1],

    "a_name": row[2],

    "b_number": row[3],

    "b_name": row[4]

}

        beacons.append(beacon)

print()
print("🐝 Beacons discovered:", len(beacons))

print()
print("🍯 First five Beacons")

for beacon in beacons[:5]:
    print(beacon)

print()
print("First five Beacons:")

for beacon in beacons[:5]:
    print(beacon)

print("🐝 Beacons discovered:", len(beacons))

print()
print("🍯 First five Beacons")

for beacon in beacons[:5]:
    print(beacon)

    search_name = "Nasty NICK"

print()
print("Searching for:", search_name)

for beacon in beacons:

    if beacon["a_name"] == search_name:

        print()
        print("🐝 Beacon Found!")
        print(beacon)

        search_name = "Nasty NICK"

print()
print("🔎 Searching for:", search_name)

for beacon in beacons:

    if beacon["a_name"] == search_name:

        print()
        print("🐝 Beacon Found!")
        print(beacon)